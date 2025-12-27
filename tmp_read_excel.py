import json
from openpyxl import load_workbook

wb = load_workbook('KODE SPAREPART.xlsx')
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
header = None
start_idx = 0
for idx, row in enumerate(rows):
    if row and any(cell is not None for cell in row):
        header = [str(c).strip() if c is not None else '' for c in row]
        start_idx = idx + 1
        break
if header is None:
    print('No data')
    raise SystemExit
records = []
for row in rows[start_idx:]:
    if row is None or all(c is None for c in row):
        continue
    rec = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
    records.append(rec)
print('Header:', header)
print('Total rows:', len(records))
print('Sample:', json.dumps(records[:20], ensure_ascii=False, indent=2))
