from pathlib import Path
import json
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
wb = load_workbook(root / "BAHAN.xlsx")

sheet_categories = {
    "Sheet1": "Produk",
    "kayu": "Kayu",
    "besi": "Besi",
}

def normalize(text: str | None) -> str:
    return (text or "").strip()

bom: dict[str, dict] = {}

# Sheet1: structured BOM (product, component, qty)
if "Sheet1" in wb.sheetnames:
    ws = wb["Sheet1"]
    for row in ws.iter_rows(values_only=True):
        product = row[0] if len(row) > 0 else None
        comp = row[1] if len(row) > 1 else None
        qty = row[2] if len(row) > 2 else None

        if isinstance(product, str) and normalize(product).upper() in {"KET", "KETERANGAN"}:
            continue
        if isinstance(comp, str) and normalize(comp).upper().startswith("RINCIAN"):
            continue

        if isinstance(product, str) and normalize(product) and isinstance(comp, str) and normalize(comp):
            try:
                q = float(qty) if qty is not None else 1.0
            except Exception:
                q = 1.0

            prod_key = normalize(product)
            entry = bom.setdefault(prod_key, {"category": "Produk", "lines": []})
            entry.setdefault("category", "Produk")
            entry.setdefault("lines", []).append({"component": normalize(comp), "qty": q})

# Kayu sheet: treat any non-empty cell as a finished good name (no BOM lines)
if "kayu" in wb.sheetnames:
    ws = wb["kayu"]
    products: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            name = normalize(cell) if isinstance(cell, str) else ""
            if name:
                products.add(name)
    for name in products:
        entry = bom.setdefault(name, {"category": "Kayu", "lines": []})
        entry["category"] = "Kayu"

# Besi sheet: use first column values as finished goods (no BOM lines)
if "besi" in wb.sheetnames:
    ws = wb["besi"]
    for row in ws.iter_rows(values_only=True):
        cell = row[0] if len(row) > 0 else None
        name = normalize(cell) if isinstance(cell, str) else ""
        if name:
            entry = bom.setdefault(name, {"category": "Besi", "lines": []})
            entry["category"] = "Besi"

out_path = root / "client" / "src" / "data" / "bom.json"
out_path.parent.mkdir(parents=True, exist_ok=True)
out_path.write_text(json.dumps(bom, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {out_path} with {len(bom)} entries across {len(sheet_categories)} sheets")
